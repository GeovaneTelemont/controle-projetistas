import json
from datetime import datetime, timedelta
from decimal import Decimal

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from django import template
from django.core.paginator import Paginator
from django.db.models import Count, Q
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.contrib import messages
from django.contrib.auth import login, authenticate, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import AuthenticationForm, PasswordChangeForm

from .forms import CustomUserCreationForm
from .models import (
    Producao,
    TipoProjeto,
    Categoria,
    Profile,
    HistoricoStatus,
    RegistroExclusao,
    User,
)


# Mantenha suas outras views como estão (custom_logout, custom_login, cadastro, dashboard, perfil)

register = template.Library()


# Adicione esta função junto com suas outras views
def exportar_excel(request):
    """
    View para exportar relatório em Excel com tempo formatado em dias, horas e tempo total
    """
    if not request.user.is_superuser:
        return HttpResponse('Acesso negado', status=403)
    
    # Obter filtros da query string
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    projetista_id = request.GET.get('projetista')
    status = request.GET.get('status')
    
    # Query base - importe o modelo Producao se necessário
    from .models import Producao  # Ajuste conforme a localização do seu modelo
    
    queryset = Producao.objects.all().select_related(
        'projetista', 'tipo_projeto', 'categoria'
    ).prefetch_related('historico')
    
    # Aplicar filtros
    if data_inicio:
        queryset = queryset.filter(data__gte=data_inicio)
    if data_fim:
        queryset = queryset.filter(data__lte=data_fim)
    if projetista_id:
        queryset = queryset.filter(projetista_id=projetista_id)
    if status:
        queryset = queryset.filter(status=status)
    
    # Ordenar por projetista e data
    queryset = queryset.order_by('projetista__username', '-data')
    
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produção por Projetista"
    
    # Definir estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    body_font = Font(name='Arial', size=10)
    body_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    

    # Função para calcular e formatar tempo
    def calcular_tempo_formatado(data_inicio, data_fim):
        """
        Retorna (dias, horas_minutos, tempo_total)
        Ex: (2, "05:30", "2 dias 05:30")
        """
        if not data_inicio or not data_fim:
            return (0, "00:00", "")
        
        # Calcular diferença em segundos
        delta = data_fim - data_inicio
        segundos_totais = delta.total_seconds()
        
        # Calcular dias inteiros
        dias = int(segundos_totais // (24 * 3600))
        
        # Calcular horas e minutos restantes
        segundos_restantes = segundos_totais % (24 * 3600)
        horas = int(segundos_restantes // 3600)
        minutos = int((segundos_restantes % 3600) // 60)
        
        # Formatar horas:minutos
        horas_minutos = f"{horas:02d}:{minutos:02d}"
        
        # Formatar tempo total
        if dias > 0:
            tempo_total = f"{dias} dia(s) {horas:02d}:{minutos:02d}"
        else:
            tempo_total = f"{horas:02d}:{minutos:02d}"
        
        return (dias, horas_minutos, tempo_total)
    
    # Cabeçalhos
    headers = [
        'DC/ID',                    # 1
        'PROJETISTAS',               # 2
        'TIPO DE PROJETO',          # 3
        'CATEGORIA',                # 4
        'STATUS',                   # 5
        'MOTIVO DE STATUS',         # 6
        'METRAGEM (M)',             # 7
        'DATA DO PROJETO',          # 8
        'DATA/HORA INÍCIO',         # 9
        'DATA/HORA TÉRMINO',        # 10
        'DIAS',                     # 11
        'HORAS',                    # 12
        'TEMPO TOTAL',              # 13
        'OBSERVAÇÕES GERAIS'        # 14
    ]
    
    # Adicionar cabeçalhos
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Preencher dados
    row_num = 2
    for producao in queryset:
        # Determinar data de término baseado no status
        data_termino = None
        
        if producao.status == 'CONCLUIDO':
            data_termino = producao.data_conclusao
        elif producao.status == 'CANCELADO':
            data_termino = producao.data_cancelamento
        
        # Calcular tempo se tiver data de término
        dias = 0
        horas_minutos = "00:00"
        tempo_total_str = ""
        
        if data_termino:
            dias, horas_minutos, tempo_total_str = calcular_tempo_formatado(
                producao.data_inicio, 
                data_termino
            )
        
        # Determinar texto da data de término
        if producao.status == 'CONCLUIDO':
            data_termino_str = producao.data_conclusao.strftime('%d/%m/%Y %H:%M') if producao.data_conclusao else ''
        elif producao.status == 'CANCELADO':
            data_termino_str = producao.data_cancelamento.strftime('%d/%m/%Y %H:%M') if producao.data_cancelamento else ''
        else:
            data_termino_str = 'EM ANDAMENTO'
        
        # Dados da linha
        row_data = [
            producao.dc_id,
            producao.projetista.get_full_name() or producao.projetista.username,
            producao.tipo_projeto.nome if producao.tipo_projeto else '',
            producao.categoria.nome if producao.categoria else '',
            producao.get_status_display(),
            producao.motivo_status or '',
            float(producao.metragem_cabo) if producao.metragem_cabo else 0.0,
            producao.data.strftime('%d/%m/%Y') if producao.data else '',
            producao.data_inicio.strftime('%d/%m/%Y %H:%M') if producao.data_inicio else '',
            data_termino_str,
            dias,
            horas_minutos,
            tempo_total_str,
            producao.observacoes or ''
        ]
        
        # Adicionar linha
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)
            cell.font = body_font
            cell.border = thin_border
            
            # Alinhamentos
            if col_num in [8, 9, 10, 11, 12, 13]:
                cell.alignment = center_alignment
            elif col_num in [5, 7]:
                cell.alignment = center_alignment
            else:
                cell.alignment = body_alignment
            
            # Formatar números
            if col_num == 7:
                cell.number_format = '#,##0.00'
            elif col_num == 11:
                cell.number_format = '0'
        
        row_num += 1
    
    # Ajustar larguras das colunas
    column_widths = {
        'A': 12, 'B': 20, 'C': 20, 'D': 15, 'E': 15,
        'F': 25, 'G': 12, 'H': 12, 'I': 18, 'J': 18,
        'K': 8, 'L': 10, 'M': 20, 'N': 40
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Congelar painel
    ws.freeze_panes = 'A2'
    
    # Criar resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'producao_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


def custom_logout(request):
    """
    View personalizada para logout com mensagem
    """
    logout(request)
    messages.success(request, 'Você foi desconectado com sucesso!')
    return redirect('dashboard')


def custom_login(request):
    """
    View personalizada para login
    """
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                messages.success(request, f'Bem-vindo, {username}!')
                
                # Redirecionar para a página que o usuário tentou acessar ou dashboard
                next_url = request.GET.get('next', 'dashboard')
                return redirect(next_url)
        else:
            messages.error(request, 'Usuário ou senha inválidos.')
    else:
        form = AuthenticationForm()
    
    return render(request, 'registration/login.html', {'form': form})


def cadastro(request):
    """
    View para cadastro de novos usuários
    """
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            
            # Fazer login automaticamente após o cadastro
            login(request, user)
            messages.success(request, f'Conta criada com sucesso! Bem-vindo, {user.username}!')
            return redirect('dashboard')
        else:
            messages.error(request, 'Por favor, corrija os erros abaixo.')
    else:
        form = CustomUserCreationForm()
    
    return render(request, 'registration/cadastro.html', {'form': form})


@login_required
def perfil(request):
    user = request.user
    profile, created = Profile.objects.get_or_create(user=user)
    
    if request.method == 'POST':
        # 1. Informações Pessoais
        if 'update_profile' in request.POST:
            user.username = request.POST.get('username', user.username)
            user.email = request.POST.get('email', user.email)
            user.first_name = request.POST.get('first_name', user.first_name)
            user.last_name = request.POST.get('last_name', user.last_name)
            user.save()
            messages.success(request, 'Informações pessoais atualizadas com sucesso!')
        
        # 2. Alterar Senha
        elif 'change_password' in request.POST:
            old_password = request.POST.get('old_password')
            new_password1 = request.POST.get('new_password1')
            new_password2 = request.POST.get('new_password2')
            
            if not user.check_password(old_password):
                messages.error(request, 'Senha atual incorreta!')
            elif new_password1 != new_password2:
                messages.error(request, 'As novas senhas não coincidem!')
            elif len(new_password1) < 8:
                messages.error(request, 'A nova senha deve ter pelo menos 8 caracteres!')
            else:
                user.set_password(new_password1)
                user.save()
                update_session_auth_hash(request, user)
                messages.success(request, 'Senha alterada com sucesso!')
        
        # 3. Foto de Perfil (upload)
        elif 'update_avatar' in request.POST and 'foto' in request.FILES:
            # Remove foto antiga se existir
            if profile.foto:
                profile.foto.delete(save=False)
            
            # Salva nova foto
            profile.foto = request.FILES['foto']
            profile.save()
            messages.success(request, 'Foto atualizada com sucesso!')
        
        # 4. Remover Foto
        elif 'remover_foto' in request.POST:
            if profile.foto:
                profile.foto.delete(save=False)
                profile.foto = None
                profile.save()
                messages.success(request, 'Foto removida com sucesso!')
        
        return redirect('perfil')
    
    context = {
        'user': user,
        'profile': profile,
    }
    return render(request, 'registration/perfil.html', context)

def dashboard(request):
    """
    View para a página inicial.
    """
    user = User.objects.all().filter(id=request.user.id).first()

    # Obter parâmetros
    data_filtro = request.GET.get('data_filtro')
    ver_todos = request.GET.get('ver_todos', 'false')  # Novo parâmetro com valor padrão
    
    # Converter para booleano
    if ver_todos.lower() == 'true':
        ver_todos = True
    else:
        ver_todos = False
    
    # Para superusuários, sempre mostrar todos
    if request.user.is_superuser:
        ver_todos = True
    
    # Base query com filtro de data se existir
    if data_filtro:
        try:
            data_filtro_obj = datetime.strptime(data_filtro, '%Y-%m-%d').date()
            if request.user.is_authenticated and not ver_todos:
                # Usuário comum vendo apenas seus projetos
                producoes = Producao.objects.filter(
                    projetista=request.user,
                    data=data_filtro_obj
                ).order_by('-data')
                base_query = Producao.objects.filter(
                    projetista=request.user,
                    data=data_filtro_obj
                )
            else:
                # Superusuário ou usuário comum escolhendo ver todos
                producoes = Producao.objects.filter(
                    data=data_filtro_obj
                ).order_by('-data')
                base_query = Producao.objects.filter(data=data_filtro_obj)
        except ValueError:
            # Data inválida, usar sem filtro
            if request.user.is_authenticated and not ver_todos:
                producoes = Producao.objects.filter(projetista=request.user).order_by('-data')
                base_query = Producao.objects.filter(projetista=request.user)
            else:
                producoes = Producao.objects.all().order_by('-data')
                base_query = Producao.objects.all()
    else:
        if request.user.is_authenticated and not ver_todos:
            producoes = Producao.objects.filter(projetista=request.user).order_by('-data')
            base_query = Producao.objects.filter(projetista=request.user)
        else:
            producoes = Producao.objects.all().order_by('-data')
            base_query = Producao.objects.all()

    # Cálculo das estatísticas
    stats_query = producoes.values('status').annotate(total=Count('id'))
    
    stats_dict = {
        'PENDENTE': 0,
        'EM_ANDAMENTO': 0,
        'REVISAO': 0,
        'CONCLUIDO': 0,
        'CANCELADO': 0,
        'total': 0
    }
    
    for stat in stats_query:
        stats_dict[stat['status']] = stat['total']
        stats_dict['total'] += stat['total']

    producoes = producoes.select_related(
        'projetista', 'tipo_projeto', 'categoria'
    ).prefetch_related('historico')

    # DADOS PARA A TABELA DE PROJETISTAS POR TIPO
    # Buscar todos os tipos de projeto
    tipos_projeto = TipoProjeto.objects.all().order_by('nome')
    
    # Buscar contagem por projetista e tipo
    projetos_por_tipo = base_query.values(
        'projetista__id',
        'projetista__username',
        'projetista__first_name',
        'projetista__last_name',
        'tipo_projeto__nome'
    ).annotate(
        count=Count('id')
    ).order_by('projetista__first_name', 'projetista__last_name')
    
    # Estruturar dados
    projetistas_data = []
    totais_por_tipo = {}
    projetistas_dict = {}

    
    for item in projetos_por_tipo:
        
        projetista_id = item['projetista__id']
        
        if projetista_id not in projetistas_dict:
            # Criar novo projetista
            nome = f"{item['projetista__first_name']} {item['projetista__last_name']}".strip()
            if not nome:
                nome = item['projetista__username']
            
            # CORREÇÃO: Buscar o usuário CORRETO pelo ID, não usar o 'user' logado
            try:
                usuario_correto = User.objects.get(id=projetista_id)  # ← MUDANÇA AQUI
            except User.DoesNotExist:
                continue  # Pula se o usuário não existir
            
            # CORREÇÃO: Pegar a foto do usuário CORRETO
            foto_url = None
            try:
                # usuario_correto.profile, não user.profile
                if hasattr(usuario_correto, 'profile') and usuario_correto.profile.foto:
                    foto_url = usuario_correto.profile.foto.url
            except Profile.DoesNotExist:
                # Cria profile se não existir
                Profile.objects.create(user=usuario_correto)
            
            projetistas_dict[projetista_id] = {
                'id': projetista_id,
                'username': item['projetista__username'],
                'email': usuario_correto.email if usuario_correto else '',  # ← MUDANÇA AQUI
                'nome': nome,
                'foto': foto_url,  # ← MUDADO de 'foto_url' para 'foto'
                'tipos': {},
                'total': 0
            }
        
        tipo_nome = item['tipo_projeto__nome']
        count = item['count']
        
        # Adicionar contagem ao tipo
        projetistas_dict[projetista_id]['tipos'][tipo_nome] = count
        projetistas_dict[projetista_id]['total'] += count
        
        # Atualizar total por tipo
        if tipo_nome in totais_por_tipo:
            totais_por_tipo[tipo_nome] += count
        else:
            totais_por_tipo[tipo_nome] = count
    
    # Converter dicionário para lista e ordenar
    projetistas_data = list(projetistas_dict.values())
    projetistas_data.sort(key=lambda x: x['total'], reverse=True)
    
    # Calcular totais
    total_geral = sum(proj['total'] for proj in projetistas_data)
    
    # Calcular média
    if projetistas_data:
        media_por_projetista = total_geral / len(projetistas_data)
    else:
        media_por_projetista = 0

    # ========== DADOS PARA O COMPONENTE ANALYTICS ==========
    
    # Período para analytics (últimos 30 dias)
    thirty_days_ago = timezone.now() - timedelta(days=30)
    
    # Base query para analytics
    if request.user.is_authenticated and not ver_todos:
        # Usuário vendo apenas seus projetos
        analytics_base_query = Producao.objects.filter(
            projetista=request.user,
            data_inicio__gte=thirty_days_ago
        )
    else:
        # Ver todos os projetos
        analytics_base_query = Producao.objects.filter(
            data_inicio__gte=thirty_days_ago
        )
    
    # Aplicar filtro de data se existir (o mesmo da tabela)
    if data_filtro:
        try:
            data_filtro_obj = datetime.strptime(data_filtro, '%Y-%m-%d').date()
            analytics_base_query = analytics_base_query.filter(data=data_filtro_obj)
        except ValueError:
            pass  # Ignora data inválida para analytics
    
    # 1. Dados para o gráfico por tipo de projeto
    project_types_data = analytics_base_query.values(
        'tipo_projeto__id', 'tipo_projeto__nome'
    ).annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Cores para o gráfico
    colors = ["#4361ee", "#3a0ca3", "#7209b7", "#f72585", "#4cc9f0", 
              "#4895ef", "#560bad", "#b5179e", "#f15bb5", "#00bbf9"]
    
    analytics_project_types = []
    for i, item in enumerate(project_types_data):
        color = colors[i % len(colors)] if i < len(colors) else f"#{i:06x}"
        analytics_project_types.append({
            'id': item['tipo_projeto__id'],
            'name': item['tipo_projeto__nome'],
            'count': item['count'],
            'color': color
        })
    
    # 2. Dados para o ranking de projetistas
    if request.user.is_authenticated and not ver_todos:
        # Usuário vendo apenas seus projetos - mostra apenas ele no ranking
        designers_query = User.objects.filter(id=request.user.id)
    else:
        # Ver todos os projetistas
        designers_query = User.objects.filter(
            producao__data_inicio__gte=thirty_days_ago
        ).distinct()
    
    # Aplicar filtro de data se existir
    if data_filtro:
        try:
            data_filtro_obj = datetime.strptime(data_filtro, '%Y-%m-%d').date()
            designers_query = designers_query.filter(
                producao__data=data_filtro_obj
            ).distinct()
        except ValueError:
            pass
    
    # Anotar estatísticas
    designers_annotated = designers_query.annotate(
        completed=Count('producao', filter=Q(
            producao__status='CONCLUIDO',
            producao__data_inicio__gte=thirty_days_ago
        )),
        in_progress=Count('producao', filter=Q(
            producao__status='EM_ANDAMENTO',
            producao__data_inicio__gte=thirty_days_ago
        )),
        total=Count('producao', filter=Q(
            producao__data_inicio__gte=thirty_days_ago
        ))
    ).filter(total__gt=0).order_by('-completed')
    
    # Processar dados dos projetistas
    analytics_designers = []
    for designer in designers_annotated:
        if designer.total > 0:
            efficiency = round((designer.completed / designer.total) * 100, 1)
        else:
            efficiency = 0
        
        # Formatar nome
        display_name = designer.get_full_name()
        if not display_name:
            display_name = designer.username
        
        analytics_designers.append({
            'id': designer.id,
            'name': display_name,
            'username': designer.username,
            'completed': designer.completed,
            'in_progress': designer.in_progress,
            'total': designer.total,
            'efficiency': efficiency
        })
    
    # 3. Estatísticas gerais para analytics
    analytics_stats = {
        'total': analytics_base_query.count(),
        'in_progress': analytics_base_query.filter(status='EM_ANDAMENTO').count(),
        'completed': analytics_base_query.filter(status='CONCLUIDO').count(),
    }
    
    # Converter dados para JSON (para usar no template mais facilmente)
    analytics_data_json = {
        'project_types': analytics_project_types,
        'designers': analytics_designers,
        'stats': analytics_stats
    }

    context = {
        'producoes': producoes,
        'stats': stats_dict,
        # Dados para a tabela
        'tipos_projeto': tipos_projeto,
        'projetistas_data': projetistas_data,
        'totais_por_tipo': totais_por_tipo,
        'total_geral': total_geral,
        'media_por_projetista': media_por_projetista,
        # Passar os filtros para o template
        'data_filtro': data_filtro,
        'ver_todos': ver_todos,
        
        # ========== DADOS PARA ANALYTICS ==========
        'analytics_project_types': analytics_project_types,
        'analytics_designers': analytics_designers,
        'analytics_stats': analytics_stats,
        'analytics_data_json': json.dumps(analytics_data_json),  # Dados em JSON
    }

    return render(request, 'dashboard.html', context)


@register.filter
def calcular_percentual(valor, total):
    """Calcula o percentual de um valor em relação ao total"""
    if total and total > 0:
        return (valor / total) * 100
    return 0

@register.filter
def get_item(dictionary, key):
    """Retorna um item de um dicionário pela chave"""
    return dictionary.get(key, 0) 


def get_client_ip(request):
    """Obtém o IP do cliente"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


def salvar_registro_exclusao(producao, request, motivo_exclusao):
    """
    Função para salvar registro de exclusão com informações da request.
    """
    # Calcular tempo total de atividade
    tempo_atividade = producao.tempo_total if hasattr(producao, 'tempo_total') else None
    
    # Contar histórico de alterações
    total_alteracoes = producao.historico.count() if hasattr(producao, 'historico') else 0
    
    # Obter informações da request
    ip_address = get_client_ip(request)
    user_agent = request.META.get('HTTP_USER_AGENT', '')
    
    # Criar registro
    RegistroExclusao.objects.create(
        dc_id=producao.dc_id,
        projeto_id_original=producao.id,
        data_projeto=producao.data,
        projetista=producao.projetista,
        tipo_projeto=producao.tipo_projeto.nome if producao.tipo_projeto else "Desconhecido",
        categoria=producao.categoria.nome if producao.categoria else "Desconhecido",
        status_final=producao.status,
        metragem_cabo=producao.metragem_cabo,
        observacoes_originais=producao.observacoes,
        total_alteracoes_status=total_alteracoes,
        tempo_total_atividade=tempo_atividade,
        motivo_exclusao=motivo_exclusao,
        usuario_exclusao=request.user,
        ip_address=ip_address,
        user_agent=user_agent
    )


@login_required
def producao(request):
    # Parâmetro para mostrar aba de inativos
    show_inativos = request.GET.get('show_inativos', False)
    
    # Produções do usuário
    producoes_queryset = Producao.objects.filter(projetista=request.user)
    
    # Separar em ativas e inativas para ordenação
    producoes_ativas = []
    producoes_inativas_recentes = []
    
    for producao_obj in producoes_queryset:
        if producao_obj.status in ['CONCLUIDO', 'CANCELADO']:
            # Verificar se foi inativado recentemente (últimas 24h)
            tempo_inativo = timezone.now() - producao_obj.updated_at
            if tempo_inativo < timedelta(days=1):
                producoes_inativas_recentes.append(producao_obj)
        else:
            producoes_ativas.append(producao_obj)
    
    # Ordenar: Ativos por data (mais recente primeiro), depois inativos recentes
    producoes_ativas.sort(key=lambda x: x.data, reverse=True)
    producoes_inativas_recentes.sort(key=lambda x: x.updated_at, reverse=True)
    
    # Combinar: ativos primeiro, depois inativos recentes
    producoes = producoes_ativas + producoes_inativas_recentes
    
    # Todas produções inativas (para aba separada)
    producoes_inativas_todas = Producao.objects.filter(
        projetista=request.user,
        status__in=['CONCLUIDO', 'CANCELADO']
    ).order_by('-updated_at')
    
    # Paginação para inativos
    page = request.GET.get('page', 1)
    paginator = Paginator(producoes_inativas_todas, 20)
    try:
        producoes_inativas_paginadas = paginator.page(page)
    except:
        producoes_inativas_paginadas = paginator.page(1)
    
    # Contadores
    producoes_concluidas = len([p for p in producoes_queryset if p.status == 'CONCLUIDO'])
    producoes_canceladas = len([p for p in producoes_queryset if p.status == 'CANCELADO'])
    producoes_andamento = len([p for p in producoes_queryset if p.status == 'EM_ANDAMENTO'])
    producoes_pendentes = len([p for p in producoes_queryset if p.status == 'PENDENTE'])
    producoes_revisao = len([p for p in producoes_queryset if p.status == 'REVISAO'])
    
    tipos_projeto = TipoProjeto.objects.all()
    categorias = Categoria.objects.all()
    status_choices = Producao.STATUS_CHOICES
    
    if request.method == 'POST':
        # ========== ADICIONAR NOVA PRODUÇÃO ==========
        if 'nova_producao' in request.POST:
            dc_id = request.POST.get('dc_id')
            data = timezone.now().date()  # Data atual automática
            tipo_projeto_id = request.POST.get('tipo_projeto')
            categoria_id = request.POST.get('categoria')
            metragem_cabo = request.POST.get('metragem_cabo', 0) or 0
            observacoes = request.POST.get('observacoes', '')
            
            
            
            try:
                producao_obj = Producao.objects.create(
                    dc_id=dc_id,
                    data=data,
                    projetista=request.user,
                    tipo_projeto_id=tipo_projeto_id,
                    categoria_id=categoria_id,
                    metragem_cabo=metragem_cabo,
                    status='EM_ANDAMENTO',
                    motivo_status='Iniciando projeto',
                    observacoes=observacoes
                )
                messages.success(request, f'Produção {dc_id} criada com sucesso!')
                return redirect('producao')
            except Exception as e:
                messages.error(request, f'Erro ao criar produção: {str(e)}')
        
        # ========== EDIÇÃO INDIVIDUAL ==========
        elif 'editar_dc_id' in request.POST:
            producao_id = request.POST.get('producao_id')
            producao_obj = get_object_or_404(Producao, id=producao_id, projetista=request.user)
            
            novo_dc_id = request.POST.get('dc_id')
            if novo_dc_id != producao_obj.dc_id and Producao.objects.filter(dc_id=novo_dc_id).exclude(id=producao_id).exists():
                messages.error(request, f'O DC/ID {novo_dc_id} já está em uso!')
            else:
                producao_obj.dc_id = novo_dc_id
                producao_obj.save()
                messages.success(request, f'DC/ID atualizado para {novo_dc_id}!')
                return redirect('producao')
        
        elif 'editar_tipo' in request.POST:
            producao_id = request.POST.get('producao_id')
            producao_obj = get_object_or_404(Producao, id=producao_id, projetista=request.user)
            
            producao_obj.tipo_projeto_id = request.POST.get('tipo_projeto')
            producao_obj.save()
            messages.success(request, 'Tipo de projeto atualizado!')
            return redirect('producao')
        
        elif 'editar_categoria' in request.POST:
            producao_id = request.POST.get('producao_id')
            producao_obj = get_object_or_404(Producao, id=producao_id, projetista=request.user)
            
            producao_obj.categoria_id = request.POST.get('categoria')
            producao_obj.save()
            messages.success(request, 'Categoria atualizada!')
            return redirect('producao')
        
        elif 'editar_metragem' in request.POST:
            producao_id = request.POST.get('producao_id')
            producao_obj = get_object_or_404(Producao, id=producao_id, projetista=request.user)
            
            producao_obj.metragem_cabo = request.POST.get('metragem_cabo', 0) or 0
            producao_obj.save()
            messages.success(request, 'Metragem atualizada!')
            return redirect('producao')
        
        elif 'editar_status' in request.POST:
            producao_id = request.POST.get('producao_id')
            producao_obj = get_object_or_404(Producao, id=producao_id, projetista=request.user)
            
            novo_status = request.POST.get('status')
            motivo = request.POST.get('motivo_status', '')
            
            # Verificar se houve mudança de status
            if novo_status != producao_obj.status:
                # Atualizar datas especiais se necessário
                if novo_status == 'CONCLUIDO' and not producao_obj.data_conclusao:
                    producao_obj.data_conclusao = timezone.now()
                elif novo_status == 'CANCELADO' and not producao_obj.data_cancelamento:
                    producao_obj.data_cancelamento = timezone.now()
                elif producao_obj.status in ['CONCLUIDO', 'CANCELADO'] and novo_status not in ['CONCLUIDO', 'CANCELADO']:
                    # Reativando - limpar datas
                    producao_obj.data_conclusao = None
                    producao_obj.data_cancelamento = None
                
                # Criar histórico
                HistoricoStatus.objects.create(
                    producao=producao_obj,
                    status=novo_status,
                    motivo=motivo,
                    usuario=request.user
                )
            
            producao_obj.status = novo_status
            producao_obj.motivo_status = motivo
            producao_obj.save()
            
            messages.success(request, f'Status atualizado para {producao_obj.get_status_display()}!')
            return redirect('producao')
        
        # ========== EDIÇÃO COMPLETA ==========
        elif 'editar_completo' in request.POST:
            producao_id = request.POST.get('producao_id')
            producao_obj = get_object_or_404(Producao, id=producao_id, projetista=request.user)
            
            dc_id = request.POST.get('dc_id')
            # Verificar se novo DC/ID já existe (se for diferente do atual)
            if dc_id != producao_obj.dc_id and Producao.objects.filter(dc_id=dc_id).exclude(id=producao_id).exists():
                messages.error(request, f'O DC/ID {dc_id} já está em uso!')
            else:
                novo_status = request.POST.get('status')
                motivo = request.POST.get('motivo_status', '')
                
                # Verificar mudança de status
                if novo_status != producao_obj.status:
                    # Atualizar datas
                    if novo_status == 'CONCLUIDO' and not producao_obj.data_conclusao:
                        producao_obj.data_conclusao = timezone.now()
                    elif novo_status == 'CANCELADO' and not producao_obj.data_cancelamento:
                        producao_obj.data_cancelamento = timezone.now()
                    elif producao_obj.status in ['CONCLUIDO', 'CANCELADO'] and novo_status not in ['CONCLUIDO', 'CANCELADO']:
                        producao_obj.data_conclusao = None
                        producao_obj.data_cancelamento = None
                    
                    # Registrar histórico
                    HistoricoStatus.objects.create(
                        producao=producao_obj,
                        status=novo_status,
                        motivo=motivo,
                        usuario=request.user
                    )
                
                producao_obj.dc_id = dc_id
                producao_obj.tipo_projeto_id = request.POST.get('tipo_projeto')
                producao_obj.categoria_id = request.POST.get('categoria')
                producao_obj.metragem_cabo = request.POST.get('metragem_cabo', 0) or 0
                producao_obj.status = novo_status
                producao_obj.motivo_status = motivo
                producao_obj.observacoes = request.POST.get('observacoes', '')
                
                producao_obj.save()
                
                messages.success(request, f'Produção {dc_id} atualizada com sucesso!')
                return redirect('producao')
        
        # ========== INATIVAR PRODUÇÃO ==========
        elif 'inativar_producao' in request.POST:
            producao_id = request.POST.get('producao_id')
            producao_obj = get_object_or_404(Producao, id=producao_id, projetista=request.user)
            
            novo_status = request.POST.get('status_inativar')
            motivo = request.POST.get('motivo_inativar', '')
            
            # Registrar histórico
            HistoricoStatus.objects.create(
                producao=producao_obj,
                status=novo_status,
                motivo=f"Inativado: {motivo}",
                usuario=request.user
            )
            
            producao_obj.status = novo_status
            producao_obj.motivo_status = motivo
            
            # Atualizar datas se necessário
            if novo_status == 'CONCLUIDO' and not producao_obj.data_conclusao:
                producao_obj.data_conclusao = timezone.now()
            elif novo_status == 'CANCELADO' and not producao_obj.data_cancelamento:
                producao_obj.data_cancelamento = timezone.now()
            
            producao_obj.save()
            
            messages.success(request, f'Produção marcada como {producao_obj.get_status_display().lower()}!')
            return redirect('producao')
        
        # ========== REATIVAR PRODUÇÃO ==========
        elif 'reativar_producao' in request.POST:
            producao_id = request.POST.get('producao_id')
            producao_obj = get_object_or_404(Producao, id=producao_id, projetista=request.user)
            
            novo_status = request.POST.get('status_reativar')
            motivo = request.POST.get('motivo_reativar', '')
            
            # Registrar histórico
            HistoricoStatus.objects.create(
                producao=producao_obj,
                status=novo_status,
                motivo=f"Reativado: {motivo}",
                usuario=request.user
            )
            
            producao_obj.status = novo_status
            producao_obj.motivo_status = motivo
            
            # Limpar datas de conclusão/cancelamento ao reativar
            if producao_obj.status not in ['CONCLUIDO', 'CANCELADO']:
                producao_obj.data_conclusao = None
                producao_obj.data_cancelamento = None
            
            producao_obj.save()
            
            messages.success(request, f'Produção reativada para {producao_obj.get_status_display().lower()}!')
            return redirect('producao')
        
        # ========== EXCLUIR PRODUÇÃO PERMANENTEMENTE ========== 
        elif 'excluir_producao' in request.POST:
            producao_id = request.POST.get('producao_id')
            producao_obj = get_object_or_404(Producao, id=producao_id, projetista=request.user)
            
            dc_id_confirmado = request.POST.get('confirmar_dc_id')
            motivo_exclusao = request.POST.get('motivo_exclusao', '')
            
            # Verificar se o DC/ID digitado confere
            if dc_id_confirmado != producao_obj.dc_id:
                messages.error(request, 'O DC/ID digitado não confere com o projeto!')
            else:
                try:
                    # Salvar registro da exclusão ANTES de excluir
                    salvar_registro_exclusao(producao_obj, request, motivo_exclusao)
                    
                    # Salvar o DC/ID para mensagem
                    dc_id = producao_obj.dc_id
                    
                    # Excluir todos os históricos associados
                    HistoricoStatus.objects.filter(producao=producao_obj).delete()
                    
                    # Excluir a produção
                    producao_obj.delete()
                    
                    messages.success(request, f'Produção {dc_id} excluída permanentemente!')
                    return redirect('producao')
                except Exception as e:
                    messages.error(request, f'Erro ao excluir produção: {str(e)}')
    
    context = {
        'producoes': producoes,  # Ativos + inativos recentes
        'producoes_inativas': producoes_inativas_paginadas,  # Paginado
        'producoes_inativas_todas': producoes_inativas_todas,  # Todos inativos
        'tipos_projeto': tipos_projeto,
        'categorias': categorias,
        'status_choices': status_choices,
        'producoes_concluidas': producoes_concluidas,
        'producoes_canceladas': producoes_canceladas,
        'producoes_andamento': producoes_andamento,
        'producoes_pendentes': producoes_pendentes,
        'producoes_revisao': producoes_revisao,
        'total_inativas': producoes_inativas_todas.count(),
        'show_inativos': show_inativos,
    }
    
    return render(request, 'producao.html', context)

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db.models import Sum, Count, Q
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .models import Producao, HistoricoStatus, TipoProjeto, Categoria

@login_required
def relatorios(request):
    """
    View única para relatórios que inclui:
    - Filtros
    - Tabela com paginação
    - Estatísticas
    - Exportação Excel
    """
    
    # ============ VERIFICAÇÃO DE EXPORTAÇÃO EXCEL ============
    if 'exportar_excel' in request.GET and request.user.is_superuser:
        return exportar_para_excel(request)
    
    # ============ FILTROS ============
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    projetista_id = request.GET.get('projetista')
    status = request.GET.get('status')
    tipo_projeto_id = request.GET.get('tipo_projeto')
    categoria_id = request.GET.get('categoria')
    
    # ============ QUERY BASE ============
    queryset = Producao.objects.all().select_related(
        'projetista', 'tipo_projeto', 'categoria'
    ).prefetch_related('historico')
    
    # ============ APLICAR FILTROS ============
    if data_inicio:
        queryset = queryset.filter(data__gte=data_inicio)
    if data_fim:
        queryset = queryset.filter(data__lte=data_fim)
    if projetista_id:
        queryset = queryset.filter(projetista_id=projetista_id)
    if status:
        queryset = queryset.filter(status=status)
    if tipo_projeto_id:
        queryset = queryset.filter(tipo_projeto_id=tipo_projeto_id)
    if categoria_id:
        queryset = queryset.filter(categoria_id=categoria_id)
    
    # ============ CALCULAR ESTATÍSTICAS ============
    total_projetos = queryset.count()
    
    # Contagem por status
    status_counts = queryset.values('status').annotate(
        count=Count('id')
    )
    
    # Converter para dicionário
    status_dict = {item['status']: item['count'] for item in status_counts}
    
    projetos_concluidos = status_dict.get('CONCLUIDO', 0)
    projetos_andamento = status_dict.get('EM_ANDAMENTO', 0)
    projetos_pendentes = status_dict.get('PENDENTE', 0)
    projetos_revisao = status_dict.get('REVISAO', 0)
    projetos_cancelados = status_dict.get('CANCELADO', 0)
    
    # Metragem total
    metragem_result = queryset.aggregate(
        total=Sum('metragem_cabo')
    )
    metragem_total = metragem_result['total'] or 0
    
    # Metragem média por projeto concluído
    if projetos_concluidos > 0:
        metragem_media = metragem_total / projetos_concluidos
    else:
        metragem_media = 0
    
    # Percentual de projetos concluídos
    percentual_concluidos = (projetos_concluidos / total_projetos * 100) if total_projetos > 0 else 0
    
    # ============ TEMPO MÉDIO - CORREÇÃO ============
    projetos_com_tempo = queryset.exclude(
        Q(data_conclusao__isnull=True) & Q(data_cancelamento__isnull=True)
    )
    
    tempo_total = timedelta()
    tempo_medio_formatado = "-"
    
    if projetos_com_tempo.exists():
        for projeto in projetos_com_tempo:
            data_final = projeto.data_conclusao or projeto.data_cancelamento
            if data_final and projeto.data_inicio:
                tempo_total += data_final - projeto.data_inicio
        
        if projetos_com_tempo.count() > 0:
            tempo_medio = tempo_total / projetos_com_tempo.count()
            
            segundos_totais = int(tempo_medio.total_seconds())
            dias = segundos_totais // 86400
            horas = (segundos_totais % 86400) // 3600
            minutos = (segundos_totais % 3600) // 60
            
            if dias > 0:
                tempo_medio_formatado = f"{dias}d {horas}h {minutos}m"
            elif horas > 0:
                tempo_medio_formatado = f"{horas}h {minutos}m"
            elif minutos > 0:
                tempo_medio_formatado = f"{minutos}m"
            else:
                tempo_medio_formatado = "< 1m"
    
    # ============ ADICIONAR CÁLCULOS DE TEMPO AOS OBJETOS ============
    for producao in queryset:
        # Determinar data de término baseado no status
        data_termino = None
        if producao.status == 'CONCLUIDO':
            data_termino = producao.data_conclusao
        elif producao.status == 'CANCELADO':
            data_termino = producao.data_cancelamento
        
        # Calcular tempo formatado usando a mesma função da exportação Excel
        dias = 0
        horas_minutos = "00:00"
        tempo_total_str = ""
        
        if data_termino and producao.data_inicio:
            delta = data_termino - producao.data_inicio
            segundos_totais = delta.total_seconds()
            
            dias = int(segundos_totais // (24 * 3600))
            segundos_restantes = segundos_totais % (24 * 3600)
            horas = int(segundos_restantes // 3600)
            minutos = int((segundos_restantes % 3600) // 60)
            
            horas_minutos = f"{horas:02d}:{minutos:02d}"
            
            if dias > 0:
                tempo_total_str = f"{dias} dia(s) {horas:02d}:{minutos:02d}"
            else:
                tempo_total_str = f"{horas:02d}:{minutos:02d}"
        else:
            # Para projetos em andamento ou pendentes
            if producao.data_inicio:
                # Calcular tempo desde o início até agora
                agora = timezone.now()
                delta = agora - producao.data_inicio
                segundos_totais = delta.total_seconds()
                
                dias = int(segundos_totais // (24 * 3600))
                segundos_restantes = segundos_totais % (24 * 3600)
                horas = int(segundos_restantes // 3600)
                minutos = int((segundos_restantes % 3600) // 60)
                
                horas_minutos = f"{horas:02d}:{minutos:02d}"
                
                if dias > 0:
                    tempo_total_str = f"{dias} dia(s) {horas:02d}:{minutos:02d}"
                else:
                    tempo_total_str = f"{horas:02d}:{minutos:02d}"
        
        # Adicionar atributos temporários ao queryset
        producao.dias_calculados = dias
        producao.horas_minutos = horas_minutos
        producao.tempo_total_formatado = tempo_total_str
        producao.data_termino_formatada = data_termino
        
        # Adicionar motivo de status (supondo que exista no modelo)
        producao.motivo_status_display = producao.motivo_status or ''
    
    # ============ ORDENAÇÃO ============
    ordenacao = request.GET.get('ordenacao', '-data')
    if ordenacao in ['data', '-data', 'projetista', '-projetista', 'dc_id', '-dc_id', 'status', '-status']:
        queryset = queryset.order_by(ordenacao)
    else:
        queryset = queryset.order_by('-data', 'projetista__username')
    
    # ============ PAGINAÇÃO ============
    itens_por_pagina = request.GET.get('itens_por_pagina', 50)
    try:
        itens_por_pagina = int(itens_por_pagina)
        if itens_por_pagina not in [10, 25, 50, 100]:
            itens_por_pagina = 50
    except:
        itens_por_pagina = 50
    
    paginator = Paginator(queryset, itens_por_pagina)
    page_number = request.GET.get('page')
    producoes = paginator.get_page(page_number)
    
    # ============ CONTEXTOS ============
    context = {
        # Dados principais
        'producoes': producoes,
        'total_projetos': total_projetos,
        'projetos_concluidos': projetos_concluidos,
        'projetos_andamento': projetos_andamento,
        'projetos_pendentes': projetos_pendentes,
        'projetos_revisao': projetos_revisao,
        'projetos_cancelados': projetos_cancelados,
        'metragem_total': metragem_total,
        'metragem_media': metragem_media,
        'percentual_concluidos': percentual_concluidos,
        
        # TEMPO MÉDIO JÁ FORMATADO
        'tempo_medio_formatado': tempo_medio_formatado,
        
        # Filtros disponíveis
        'projetistas': User.objects.filter(
            groups__name='Projetistas'
        ).order_by('first_name') if User.objects.filter(groups__name='Projetistas').exists() 
          else User.objects.filter(is_active=True).order_by('first_name'),
        'tipos_projeto': TipoProjeto.objects.all().order_by('nome'),
        'categorias': Categoria.objects.all().order_by('nome'),
        'status_choices': Producao.STATUS_CHOICES,
        
        # Valores atuais dos filtros
        'filter_data': {
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'projetista': projetista_id,
            'status': status,
            'tipo_projeto': tipo_projeto_id,
            'categoria': categoria_id,
            'ordenacao': ordenacao,
            'itens_por_pagina': itens_por_pagina,
        },
        
        # Permissões
        'is_superuser': request.user.is_superuser,
        
        # Para paginação
        'paginator': paginator,
    }
    
    return render(request, 'relatorios.html', context)


def exportar_para_excel(request):
    """
    Função interna para exportar dados para Excel
    """
    if not request.user.is_superuser:
        return HttpResponse('Acesso negado', status=403)
    
    # Obter filtros da query string
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    projetista_id = request.GET.get('projetista')
    status = request.GET.get('status')
    tipo_projeto_id = request.GET.get('tipo_projeto')
    categoria_id = request.GET.get('categoria')
    
    # Query base
    queryset = Producao.objects.all().select_related(
        'projetista', 'tipo_projeto', 'categoria'
    ).prefetch_related('historico')
    
    # Aplicar filtros
    if data_inicio:
        queryset = queryset.filter(data__gte=data_inicio)
    if data_fim:
        queryset = queryset.filter(data__lte=data_fim)
    if projetista_id:
        queryset = queryset.filter(projetista_id=projetista_id)
    if status:
        queryset = queryset.filter(status=status)
    if tipo_projeto_id:
        queryset = queryset.filter(tipo_projeto_id=tipo_projeto_id)
    if categoria_id:
        queryset = queryset.filter(categoria_id=categoria_id)
    
    # Ordenar por projetista e data
    queryset = queryset.order_by('projetista__username', '-data')
    
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produção por Projetista"
    
    # Definir estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    body_font = Font(name='Arial', size=10)
    body_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Função para calcular e formatar tempo
    def calcular_tempo_formatado(data_inicio, data_fim):
        """
        Retorna (dias, horas_minutos, tempo_total)
        """
        if not data_inicio or not data_fim:
            return (0, "00:00", "")
        
        # Calcular diferença em segundos
        delta = data_fim - data_inicio
        segundos_totais = delta.total_seconds()
        
        # Calcular dias inteiros
        dias = int(segundos_totais // (24 * 3600))
        
        # Calcular horas e minutos restantes
        segundos_restantes = segundos_totais % (24 * 3600)
        horas = int(segundos_restantes // 3600)
        minutos = int((segundos_restantes % 3600) // 60)
        
        # Formatar horas:minutos
        horas_minutos = f"{horas:02d}:{minutos:02d}"
        
        # Formatar tempo total
        if dias > 0:
            tempo_total = f"{dias} dia(s) {horas:02d}:{minutos:02d}"
        else:
            tempo_total = f"{horas:02d}:{minutos:02d}"
        
        return (dias, horas_minutos, tempo_total)
    
    # Cabeçalhos
    headers = [
        'DC/ID',                    # 1
        'Projetista',               # 2
        'Tipo de Projeto',          # 3
        'Categoria',                # 4
        'Status',                   # 5
        'Motivo do Status',         # 6
        'Metragem (m)',             # 7
        'Data do Projeto',          # 8
        'Data/Hora Início',         # 9
        'Data/Hora Término',        # 10
        'DIAS',                     # 11
        'HORAS',                    # 12
        'TEMPO TOTAL',              # 13
        'Observações Gerais',       # 14
        'Histórico de Status'       # 15 - Novo campo
    ]
    
    # Adicionar cabeçalhos
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Preencher dados
    row_num = 2
    for producao in queryset:
        # Determinar data de término baseado no status
        data_termino = None
        
        if producao.status == 'CONCLUIDO':
            data_termino = producao.data_conclusao
        elif producao.status == 'CANCELADO':
            data_termino = producao.data_cancelamento
        
        # Calcular tempo se tiver data de término
        dias = 0
        horas_minutos = "00:00"
        tempo_total_str = ""
        
        if data_termino:
            dias, horas_minutos, tempo_total_str = calcular_tempo_formatado(
                producao.data_inicio, 
                data_termino
            )
        
        # Determinar texto da data de término
        if producao.status == 'CONCLUIDO':
            data_termino_str = producao.data_conclusao.strftime('%d/%m/%Y %H:%M') if producao.data_conclusao else ''
        elif producao.status == 'CANCELADO':
            data_termino_str = producao.data_cancelamento.strftime('%d/%m/%Y %H:%M') if producao.data_cancelamento else ''
        else:
            data_termino_str = 'EM ANDAMENTO'
        
        # Obter histórico formatado
        historico = producao.historico.all().order_by('data_alteracao')
        historico_texto = ""
        for h in historico:
            data_formatada = h.data_alteracao.strftime('%d/%m/%Y %H:%M')
            usuario = h.usuario.get_full_name() if h.usuario and h.usuario.get_full_name() else (h.usuario.username if h.usuario else 'Sistema')
            historico_texto += f"{data_formatada} - {h.get_status_display()} - {usuario}"
            if h.motivo:
                historico_texto += f" | Motivo: {h.motivo}"
            historico_texto += "\n"
        
        # Dados da linha
        row_data = [
            producao.dc_id,
            producao.projetista.get_full_name() or producao.projetista.username,
            producao.tipo_projeto.nome if producao.tipo_projeto else '',
            producao.categoria.nome if producao.categoria else '',
            producao.get_status_display(),
            producao.motivo_status or '',
            float(producao.metragem_cabo) if producao.metragem_cabo else 0.0,
            producao.data.strftime('%d/%m/%Y') if producao.data else '',
            producao.data_inicio.strftime('%d/%m/%Y %H:%M') if producao.data_inicio else '',
            data_termino_str,
            dias,
            horas_minutos,
            tempo_total_str,
            producao.observacoes or '',
            historico_texto.strip()
        ]
        
        # Adicionar linha
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)
            cell.font = body_font
            cell.border = thin_border
            
            # Alinhamentos
            if col_num in [8, 9, 10, 11, 12, 13]:  # Datas e tempos
                cell.alignment = center_alignment
            elif col_num in [5, 7]:  # Status e Metragem
                cell.alignment = center_alignment
            elif col_num == 15:  # Histórico (multi-linha)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:
                cell.alignment = body_alignment
            
            # Formatar números
            if col_num == 7:
                cell.number_format = '#,##0.00'
            elif col_num == 11:
                cell.number_format = '0'
        
        row_num += 1
    
    # Adicionar uma aba com resumo
    ws2 = wb.create_sheet(title="Resumo")
    
    # Cabeçalho do resumo
    ws2['A1'] = "RESUMO DE PRODUÇÃO"
    ws2['A1'].font = Font(name='Arial', size=14, bold=True)
    
    # Dados do resumo
    resumo_data = [
        ["Período:", f"{data_inicio or 'Início'} a {data_fim or 'Fim'}"],
        ["Data de Exportação:", datetime.now().strftime('%d/%m/%Y %H:%M')],
        ["Exportado por:", request.user.get_full_name() or request.user.username],
        ["", ""],
        ["TOTAL DE PROJETOS:", total_projetos],
        ["Projetos Concluídos:", projetos_concluidos],
        ["Projetos em Andamento:", projetos_andamento],
        ["Projetos Pendentes:", projetos_pendentes],
        ["Projetos em Revisão:", projetos_revisao],
        ["Projetos Cancelados:", projetos_cancelados],
        ["", ""],
        ["METRAGEM TOTAL:", f"{metragem_total:,.2f} m"],
        ["TEMPO MÉDIO POR PROJETO:", str(tempo_medio).split('.')[0]]
    ]
    
    # Preencher resumo
    for i, (label, value) in enumerate(resumo_data, start=3):
        ws2.cell(row=i, column=1, value=label)
        ws2.cell(row=i, column=1).font = Font(name='Arial', size=11, bold=True)
        
        ws2.cell(row=i, column=2, value=value)
        if i >= 8 and i <= 12:  # Valores numéricos
            ws2.cell(row=i, column=2).number_format = '#,##0'
    
    # Ajustar larguras das colunas na primeira aba
    column_widths = {
        'A': 12, 'B': 20, 'C': 20, 'D': 15, 'E': 15,
        'F': 25, 'G': 12, 'H': 12, 'I': 18, 'J': 18,
        'K': 8, 'L': 10, 'M': 20, 'N': 40, 'O': 60
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Ajustar larguras na segunda aba
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 30
    
    # Congelar painel na primeira linha
    ws.freeze_panes = 'A2'
    
    # Criar resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'producao_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

